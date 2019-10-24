from openpyxl import load_workbook
# from openpyxl.styles import colors, Font, Fill, NamedStyle
# from openpyxl.styles import PatternFill, Border, Side, Alignment

import json
import csv

# import records
import pymysql


class BomItem:
    def __init__(self, bom_id, seq_num, level, part_num, part_version, part_name, part_family, part_name_zh):
        self.bom_id = bom_id
        self.seq_num = seq_num
        self.level = level
        self.part_num = part_num
        self.part_version = part_version
        self.part_name = part_name
        self.part_family = part_family
        self.part_name_zh = part_name_zh

    def dump(self):
        return json.dumps(self.__dict__)


boms = []


def save_file(items):
    csv_columns = ['bom_id', 'seq_num', 'level', 'part_num',
                   'part_version', 'part_name', 'part_family', 'part_name_zh']
    with open('test.csv', 'w') as f:
        writer = csv.DictWriter(f, fieldnames=csv_columns, delimiter="|")
        writer.writeheader()
        for i in items:
            writer.writerow(i.__dict__)


def purify(s):
    return s.strip().replace('\r', ' ').replace('\n', ' ').replace('"', '').replace("'", '')


def purify_version(s):
    a = purify(s)
    return a.replace(',', '').replace('，', '').strip()


def parse_C490_front():
    # 加载文件
    fname = '/Users/ygzheng/Documents/project/Support/延锋/07.BOM/BOM/parse/c490.xlsx'
    wb = load_workbook(fname)

    # 读取sheetname
    print('输出文件所有工作表名：\n', wb.sheetnames)

    # 或者不知道名字时
    # sheet_names = wb.sheetnames   # 返回一个列表
    # ws = wb[sheet_names[0]]    # index为0为第一张表
    ws = wb['C490-Front']

    # 循环每一行
    for r in range(4, 500):
        seq_num = ws.cell(row=r, column=1).value
        if (not seq_num):
            break

        level = ws.cell(row=r, column=2).value

        part_name = ''
        for c in range(3, 13):
            n = ws.cell(row=r, column=c).value
            if n:
                part_name = n
                break

        part_num = ws.cell(row=r, column=13).value
        part_version = ws.cell(row=r, column=14).value
        part_family = ws.cell(row=r, column=15).value

        i = BomItem('C490-Front', seq_num, level, part_num, part_version,
                    part_name, part_family, '')

        boms.append(i)


def parse_C490_rear():
    # 加载文件
    fname = '/Users/ygzheng/Documents/project/Support/延锋/07.BOM/BOM/parse/c490.xlsx'
    wb = load_workbook(fname)

    # 读取sheetname
    print('输出文件所有工作表名：\n', wb.sheetnames)

    ws = wb['C490-Rear']

    # 循环每一行
    for r in range(4, 500):
        seq_num = ws.cell(row=r, column=1).value
        if (not seq_num):
            break

        level = ws.cell(row=r, column=2).value

        part_name = ''
        for c in range(3, 13):
            n = ws.cell(row=r, column=c).value
            if n:
                part_name = n
                break

        part_name_zh = ws.cell(row=r, column=13).value
        part_num = ws.cell(row=r, column=15).value
        part_version = ws.cell(row=r, column=16).value
        part_family = ws.cell(row=r, column=17).value

        i = BomItem('C490-Rear', seq_num, level, part_num, part_version,
                    part_name, part_family, part_name_zh)

        boms.append(i)


def parse_C519():
    # 加载文件
    fname = '/Users/ygzheng/Documents/project/Support/延锋/07.BOM/BOM/parse/c519.xlsm'
    wb = load_workbook(fname)

    # 读取sheetname
    print('输出文件所有工作表名：\n', wb.sheetnames)

    ws = wb['C519']

    # 循环每一行
    for r in range(7, 700):
        seq_num = ws.cell(row=r, column=1).value
        if (not seq_num):
            break

        level = 99
        for c in range(2, 12):
            n = ws.cell(row=r, column=c).value
            if n:
                level = n
                break

        part_num = ws.cell(row=r, column=12).value
        if (not part_num):
            continue

        if isinstance(part_num, str):
            part_num = part_num.strip()

        part_version = ws.cell(row=r, column=13).value
        if (part_version and isinstance(part_version, str)):
            part_version = part_version.strip()

        part_name = ws.cell(row=r, column=26).value
        if part_name:
            part_name = part_name.strip()
            part_name = part_name.replace('\n', ' ').replace('\r', ' ')

        part_name_zh = ws.cell(row=r, column=27).value
        if part_name_zh:
            part_name_zh = part_name_zh.strip()

        part_family = ws.cell(row=r, column=29).value

        i = BomItem('C519', seq_num, level, part_num, part_version,
                    part_name, part_family, part_name_zh)

        boms.append(i)

        # print(i.dump())


def parse_JH476():
    # 加载文件
    fname = '/Users/ygzheng/Documents/project/Support/延锋/07.BOM/BOM/parse/c519.xlsm'
    wb = load_workbook(fname)

    # 读取sheetname
    print('输出文件所有工作表名：\n', wb.sheetnames)

    ws = wb['JH476']

    # 循环每一行
    for r in range(7, 700):
        seq_num = ws.cell(row=r, column=1).value

        level = 99
        for c in range(2, 12):
            n = ws.cell(row=r, column=c).value
            if n:
                level = n
                break

        part_num = ws.cell(row=r, column=12).value
        if ((not seq_num) and (not part_num)):
            break

        if ((seq_num) and (not part_num)):
            continue

        if isinstance(part_num, str):
            part_num = part_num.strip()

        part_version = ws.cell(row=r, column=13).value
        if (part_version and isinstance(part_version, str)):
            part_version = part_version.strip()

        part_name = ws.cell(row=r, column=24).value
        if part_name:
            part_name = part_name.strip().replace('\n', ' ').replace('\r', ' ')

        part_name_zh = ws.cell(row=r, column=25).value
        if part_name_zh:
            part_name_zh = part_name_zh.strip().replace('\n', ' ').replace('\r', ' ')

        part_family = ws.cell(row=r, column=27).value

        i = BomItem('JH476', seq_num, level, part_num, part_version,
                    part_name, part_family, part_name_zh)

        boms.append(i)

        # print(i.dump())


def parse_nanchang():
    # 加载文件
    fname = '/Users/ygzheng/Documents/project/Support/延锋/07.BOM/BOM/parse/nanchang.xlsm'
    wb = load_workbook(fname)

    # 读取sheetname
    print('输出文件所有工作表名：\n', wb.sheetnames)

    # 列的位置完全一样
    tabs = ['M11-FS', 'M11-RS5', 'M11-MS7', 'M11-RS7', 'M11-Trim-5',
            'M11-Trim-7', 'PSA05-FS', 'PSA05-RS', 'PSA05-MS']
    # tabs = ['PSA05-MS']
    for t in tabs:
        print('{}: start'.format(t))
        ws = wb[t]

        # 循环每一行
        for r in range(7, 700):
            seq_num = ws.cell(row=r, column=1).value

            level = 99
            for c in range(2, 12):
                n = ws.cell(row=r, column=c).value
                if n:
                    level = n
                    break

            part_num = ws.cell(row=r, column=12).value
            if ((not seq_num) and (not part_num)):
                break

            if ((seq_num) and (not part_num)):
                continue

            if isinstance(part_num, str):
                part_num = purify(part_num)

            part_version = ws.cell(row=r, column=13).value
            if (part_version and isinstance(part_version, str)):
                part_version = purify_version(part_version)

            part_name = ws.cell(row=r, column=24).value
            if part_name:
                part_name = purify(part_name)

            part_name_zh = ws.cell(row=r, column=25).value
            if part_name_zh:
                part_name_zh = purify(part_name_zh)

            part_family = ws.cell(row=r, column=27).value

            i = BomItem(t, seq_num, level, part_num, part_version,
                        part_name, part_family, part_name_zh)

            boms.append(i)

        print('{}: end'.format(t))

        # print(i.dump())


def parse_nanchang_U375():
    # 加载文件
    fname = '/Users/ygzheng/Documents/project/Support/延锋/07.BOM/BOM/parse/nanchang_U375.xlsm'
    wb = load_workbook(fname)

    # 读取sheetname
    print('输出文件所有工作表名：\n', wb.sheetnames)

    # 列的位置完全一样
    tabs = ['U375_1RW', 'U375_2RW', 'U375_2RW_China',
            'U375_3RW', 'U375_3RW_China']
    # tabs = ['U375_3RW_China']
    for t in tabs:
        print('{}: start'.format(t))

        ws = wb[t]

        # 循环每一行
        for r in range(3, 700):
            seq_num = ws.cell(row=r, column=1).value
            if (not seq_num):
                break

            level = ws.cell(row=r, column=2).value

            part_name = ''
            for c in range(3, 13):
                n = ws.cell(row=r, column=c).value
                if n:
                    part_name = purify(n)
                    break

            part_name_zh = ws.cell(row=r, column=13).value
            part_num = ws.cell(row=r, column=14).value
            part_version = ws.cell(row=r, column=15).value
            part_family = ws.cell(row=r, column=16).value

            i = BomItem(t, seq_num, level, part_num, part_version,
                        part_name, part_family, part_name_zh)

            boms.append(i)

        print('{}: end'.format(t))


def parse_nanchang_V362():
    # 加载文件
    fname = '/Users/ygzheng/Documents/project/Support/延锋/07.BOM/BOM/parse/nanchang_V362.xlsm'
    wb = load_workbook(fname)

    # 读取sheetname
    print('输出文件所有工作表名：\n', wb.sheetnames)

    # 列的位置完全一样
    tabs = ['V362C_2Row_43', 'V362c_2Row_46', 'V362C_Kombi', 'V362C_Flamingo', 'V362_DualCo',
            'V362_Front_pedestal', 'V362_Front_10w', 'V362_Front_6w', 'V362_Front_2w']
    # tabs = ['V362C_2Row_43']
    for t in tabs:
        print('{}: start'.format(t))

        ws = wb[t]

        # 循环每一行
        for r in range(5, 700):
            idx = r - 4

            seq_num = ws.cell(row=r, column=1).value

            level = 99
            for c in range(2, 15):
                n = ws.cell(row=r, column=c).value
                if n:
                    level = n
                    break

            part_num = ws.cell(row=r, column=15).value
            if ((not seq_num) and (not part_num)):
                break

            if ((seq_num) and (not part_num)):
                continue

            if isinstance(part_num, str):
                part_num = purify(part_num)

            part_version = ws.cell(row=r, column=16).value
            if (part_version and isinstance(part_version, str)):
                part_version = purify_version(part_version)

            part_family = ws.cell(row=r, column=17).value

            part_name = ws.cell(row=r, column=18).value
            if part_name:
                part_name = purify(part_name)

            part_name_zh = ws.cell(row=r, column=19).value
            if part_name_zh:
                part_name_zh = purify(part_name_zh)

            i = BomItem(t, idx, level, part_num, part_version,
                        part_name, part_family, part_name_zh)

            boms.append(i)

        print('{}: end'.format(t))


def parse_taizhou():
    # 加载文件
    fname = '/Users/ygzheng/Documents/project/Support/延锋/07.BOM/BOM/parse/taizhou.xlsm'
    wb = load_workbook(fname)

    # 读取sheetname
    print('输出文件所有工作表名：\n', wb.sheetnames)

    # 列的位置完全一样
    tabs = ['FS11', 'K316_FS', 'K316_RS', 'WM_Ajax', 'WM_ASE2_FS',
            'WM_ASE2_RS_6_3', 'WM_ASE2_RS_6_2', 'WM_ASE2_RS_5']
    # tabs = ['FS11']
    for t in tabs:
        print('{}: start'.format(t))

        ws = wb[t]

        # 循环每一行
        for r in range(7, 700):
            idx = r - 6

            seq_num = ws.cell(row=r, column=1).value

            level = 99
            for c in range(2, 15):
                n = ws.cell(row=r, column=c).value
                if n:
                    level = n
                    break

            part_num = ws.cell(row=r, column=15).value
            if ((not seq_num) and (not part_num)):
                break

            if ((seq_num) and (not part_num)):
                continue

            if isinstance(part_num, str):
                part_num = purify(part_num)

            part_version = ws.cell(row=r, column=16).value
            if (part_version and isinstance(part_version, str)):
                part_version = purify_version(part_version)

            part_family = ws.cell(row=r, column=17).value

            part_name = ws.cell(row=r, column=18).value
            if part_name:
                part_name = purify(part_name)

            part_name_zh = ws.cell(row=r, column=19).value
            if part_name_zh:
                part_name_zh = purify(part_name_zh)

            i = BomItem(t, idx, level, part_num, part_version,
                        part_name, part_family, part_name_zh)

            boms.append(i)

        print('{}: end'.format(t))


def parse_hefei():
    # 加载文件
    fname = '/Users/ygzheng/Documents/project/Support/延锋/07.BOM/BOM/parse/hefei.xlsm'
    wb = load_workbook(fname)

    # 读取sheetname
    print('输出文件所有工作表名：\n', wb.sheetnames)

    # 列的位置完全一样
    tabs = ['ES6_FS', 'ES6_RS', 'ES8_FS', 'ES8_MS', 'ES8_RS', 'L541_RS']
    # tabs = ['ES6_FS']
    for t in tabs:
        print('{}: start'.format(t))

        ws = wb[t]

        # 循环每一行
        for r in range(7, 1600):
            idx = r - 6

            seq_num = ws.cell(row=r, column=1).value
            if seq_num == 'ENDEND':
                break

            level = 99
            for c in range(2, 15):
                n = ws.cell(row=r, column=c).value
                if n is not None:
                    level = n
                    break

            part_num = ws.cell(row=r, column=15).value
            if (not part_num):
                continue

            if isinstance(part_num, str):
                part_num = purify(part_num)

            part_version = ws.cell(row=r, column=16).value
            if (part_version and isinstance(part_version, str)):
                part_version = purify_version(part_version)

            part_family = ws.cell(row=r, column=17).value

            part_name = ws.cell(row=r, column=18).value
            if part_name:
                part_name = purify(part_name)

            part_name_zh = ws.cell(row=r, column=19).value
            if part_name_zh:
                part_name_zh = purify(part_name_zh)

            i = BomItem(t, idx, level, part_num, part_version,
                        part_name, part_family, part_name_zh)

            boms.append(i)

        print('{}: end'.format(t))


def parse_chengdu():
    # 加载文件
    fname = '/Users/ygzheng/Documents/project/Support/延锋/07.BOM/BOM/parse/chengdu.xlsm'
    wb = load_workbook(fname)

    # 读取sheetname
    print('输出文件所有工作表名：\n', wb.sheetnames)

    # 列的位置完全一样
    tabs = ['K426_FS', 'K426_RS', 'K426_518_FS',
            'K426_518_RS', 'NL_4AB_FS', 'NL_4AB_RS']
    # tabs = ['NL_4AB_FS']
    for t in tabs:
        print('{}: start'.format(t))

        ws = wb[t]

        # 循环每一行
        for r in range(7, 1100):
            idx = r - 6

            seq_num = ws.cell(row=r, column=1).value
            if seq_num == 'ENDEND':
                break

            level = 99
            for c in range(2, 15):
                n = ws.cell(row=r, column=c).value
                if n is not None:
                    level = n
                    break

            part_num = ws.cell(row=r, column=15).value
            if (not part_num):
                continue

            if isinstance(part_num, str):
                part_num = purify(part_num)

            part_version = ws.cell(row=r, column=16).value
            if (part_version and isinstance(part_version, str)):
                part_version = purify_version(part_version)

            part_family = ws.cell(row=r, column=17).value

            part_name = ws.cell(row=r, column=18).value
            if part_name:
                part_name = purify(str(part_name))

            part_name_zh = ws.cell(row=r, column=19).value
            if part_name_zh:
                part_name_zh = purify(part_name_zh)

            i = BomItem(t, idx, level, part_num, part_version,
                        part_name, part_family, part_name_zh)

            boms.append(i)

        print('{}: end'.format(t))


def parse_hangzhou():
    # 加载文件
    fname = '/Users/ygzheng/Documents/project/Support/延锋/07.BOM/BOM/parse/hangzhou.xlsm'
    wb = load_workbook(fname)

    # 读取sheetname
    print('输出文件所有工作表名：\n', wb.sheetnames)

    # 列的位置完全一样
    tabs = ['U611_FS', 'U611_MS', 'U611_RS', 'U625_FS', 'U625_MS', 'U625_RS']
    # tabs = ['U611_FS']
    for t in tabs:
        print('{}: start'.format(t))

        ws = wb[t]

        # 循环每一行
        for r in range(7, 500):
            idx = r - 6

            seq_num = ws.cell(row=r, column=1).value
            if seq_num == 'ENDEND':
                break

            level = 99
            for c in range(2, 15):
                n = ws.cell(row=r, column=c).value
                if n is not None:
                    level = n
                    break

            part_num = ws.cell(row=r, column=15).value
            if (not part_num):
                continue

            if isinstance(part_num, str):
                part_num = purify(part_num)

            part_version = ws.cell(row=r, column=16).value
            if (part_version and isinstance(part_version, str)):
                part_version = purify_version(part_version)

            part_family = ws.cell(row=r, column=17).value

            part_name = ws.cell(row=r, column=18).value
            if part_name:
                part_name = purify(str(part_name))

            part_name_zh = ws.cell(row=r, column=19).value
            if part_name_zh:
                part_name_zh = purify(part_name_zh)

            i = BomItem(t, idx, level, part_num, part_version,
                        part_name, part_family, part_name_zh)

            boms.append(i)

        print('{}: end'.format(t))


def parse_jiangjin():
    # 加载文件
    fname = '/Users/ygzheng/Documents/project/Support/延锋/07.BOM/BOM/parse/jiangjin.xlsm'
    wb = load_workbook(fname)

    # 读取sheetname
    print('输出文件所有工作表名：\n', wb.sheetnames)

    # 列的位置完全一样
    # tabs = ['YAE']
    tabs = ['YAE', 'YAE_TRIM_17', 'YFE', 'YFE_TRIM', 'YL1_FS', 'YL1_RS', 'YL1_Trim', 'M12_FS', 'M12_RS', 'F506S_FS', 'F506S_MS',
            'F506S_RS', 'F507S_FS', 'F507S_MS', 'F507_RS', 'F516_FS', 'F516_MS', 'F516_RS', 'F517_FS', 'F517_RS', 'F517_MS']
    for t in tabs:
        print('{}: start'.format(t))

        ws = wb[t]

        # 循环每一行
        for r in range(7, 1500):
            idx = r - 6

            seq_num = ws.cell(row=r, column=1).value
            if seq_num == 'ENDEND':
                break

            level = 99
            for c in range(2, 15):
                n = ws.cell(row=r, column=c).value
                if n is not None:
                    level = n
                    break

            part_num = ws.cell(row=r, column=15).value
            if (not part_num):
                continue

            if isinstance(part_num, str):
                part_num = purify(part_num)

            part_version = ws.cell(row=r, column=16).value
            if (part_version and isinstance(part_version, str)):
                part_version = purify_version(part_version)

            part_family = ws.cell(row=r, column=17).value

            part_name = ws.cell(row=r, column=18).value
            if part_name:
                part_name = purify(str(part_name))

            part_name_zh = ws.cell(row=r, column=19).value
            if part_name_zh:
                part_name_zh = purify(part_name_zh)

            i = BomItem(t, idx, level, part_num, part_version,
                        part_name, part_family, part_name_zh)

            boms.append(i)

        print('{}: end'.format(t))


def parse_yuzui():
    # 加载文件
    fname = '/Users/ygzheng/Documents/project/Support/延锋/07.BOM/BOM/parse/yuzui.xlsm'
    wb = load_workbook(fname)

    # 读取sheetname
    print('输出文件所有工作表名：\n', wb.sheetnames)

    # 列的位置完全一样
    # tabs = ['YAE']
    tabs = ['CN115_FS_LH', 'CN115_FS_RH', 'CN115_MS_LH', 'CN115_MS_RH', 'CN115_RS', 'C301',
            'S201_DRV', 'S201_PSG', 'S201_RS', 'S201_Lock', 'S201_Trim', 'S301_Trim', 'S402_FS ']
    for t in tabs:
        print('{}: start'.format(t))

        ws = wb[t]

        # 循环每一行
        for r in range(7, 1500):
            idx = r - 6

            seq_num = ws.cell(row=r, column=1).value
            if seq_num == 'ENDEND':
                break

            level = 99
            for c in range(2, 15):
                n = ws.cell(row=r, column=c).value
                if n is not None:
                    level = n
                    break

            part_num = ws.cell(row=r, column=15).value
            if (not part_num):
                continue

            if isinstance(part_num, str):
                part_num = purify(part_num)

            part_version = ws.cell(row=r, column=16).value
            if (part_version and isinstance(part_version, str)):
                part_version = purify_version(part_version)

            part_family = ws.cell(row=r, column=17).value

            part_name = ws.cell(row=r, column=18).value
            if part_name:
                part_name = purify(str(part_name))

            part_name_zh = ws.cell(row=r, column=19).value
            if part_name_zh:
                part_name_zh = purify(part_name_zh)

            i = BomItem(t, idx, level, part_num, part_version,
                        part_name, part_family, part_name_zh)

            boms.append(i)

        print('{}: end'.format(t))


def parse_dongguan():
    # 加载文件
    fname = '/Users/ygzheng/Documents/project/Support/延锋/07.BOM/BOM/parse/dongguan.xlsm'
    wb = load_workbook(fname)

    # 读取sheetname
    print('输出文件所有工作表名：\n', wb.sheetnames)

    # 列的位置完全一样
    tabs = ['X74_FS', 'X74_RS', 'XiaoPeng_F1', 'XiaoPeng_F2', 'XiaoPeng_RS']
    for t in tabs:
        print('{}: start'.format(t))

        ws = wb[t]

        # 循环每一行
        for r in range(7, 500):
            idx = r - 6

            seq_num = ws.cell(row=r, column=1).value
            if seq_num == 'ENDEND':
                break

            level = 99
            for c in range(2, 15):
                n = ws.cell(row=r, column=c).value
                if n is not None:
                    level = n
                    break

            part_num = ws.cell(row=r, column=15).value
            if (not part_num):
                continue

            if isinstance(part_num, str):
                part_num = purify(part_num)

            part_version = ws.cell(row=r, column=16).value
            if (part_version and isinstance(part_version, str)):
                part_version = purify_version(part_version)

            part_family = ws.cell(row=r, column=17).value

            part_name = ws.cell(row=r, column=18).value
            if part_name:
                part_name = purify(str(part_name))

            part_name_zh = ws.cell(row=r, column=19).value
            if part_name_zh:
                part_name_zh = purify(part_name_zh)

            i = BomItem(t, idx, level, part_num, part_version,
                        part_name, part_family, part_name_zh)

            boms.append(i)

        print('{}: end'.format(t))


def parse_all():
    # parse_C490_front()
    # parse_C490_rear()
    # parse_C519()
    # parse_JH476()

    # parse_nanchang()
    # parse_nanchang_U375()
    # parse_nanchang_V362()

    # parse_taizhou()
    # parse_hefei()
    parse_chengdu()
    # parse_hangzhou()
    # parse_jiangjin()
    # parse_yuzui()
    # parse_dongguan()

    save_file(boms)


all_items = []


def load_file():
    csv_columns = ['bom_id', 'seq_num', 'level', 'part_num',
                   'part_version', 'part_name', 'part_family', 'part_name_zh']

    fs = ['HMP.csv', 'chengdu.csv', 'dongguan.csv', 'hangzhou.csv',
          'hefei.csv', 'jiangjin.csv', 'nanchang1.csv', 'nanchang_U375.csv', 'nanchang_V362.csv', 'taizhou.csv',
          'yuzui.csv']

    # fs = ['ztest.csv']
    for f in fs:
        fname = './bom/data/{}'.format(f)
        print(fname)

        with open(fname, mode='r') as csv_file:
            csv_reader = csv.DictReader(
                csv_file, fieldnames=csv_columns, delimiter="|")

            line_count = 0
            for row in csv_reader:
                if line_count == 0:
                    line_count += 1
                    continue

                line_count += 1

                i = BomItem(row['bom_id'], row['seq_num'], row['level'], row['part_num'],
                            row['part_version'], row['part_name'], row['part_family'], row['part_name_zh'])

                all_items.append(i)

    print('total: {} lines.'.format(len(all_items)))


def tryMysql():
    # 连接数据库
    conn = pymysql.connect(host='127.0.0.1', user='root',
                           password='mysql', database='cqyfas', port=3333, charset='utf8')
    cur = conn.cursor()

    version = "SELECT VERSION()"
    cur.execute(version)
    data = cur.fetchone()
    print("Database version : %s " % data)

    load_file()

    i = 0
    for row in all_items:
        i += 1
        # cmd = "insert into bom_items (bom_id, seq_num, level, part_num, part_version, part_name, part_family, part_name_zh) values ('{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}')".format(row.bom_id, row.seq_num, row.level, row.part_num,
        #                                                                                                                                                                                     row.part_version, row.part_name, row.part_family, row.part_name_zh)
        # print(cmd)
        # cur.execute(cmd)

        cmd = 'insert into bom_items (bom_id, seq_num, level, part_num, part_version, part_name, part_family, part_name_zh) values (%s, %s,%s,%s,%s,%s,%s,%s)'
        cur.execute(cmd, (row.bom_id, row.seq_num, row.level, row.part_num, row.part_version, row.part_name, row.part_family, row.part_name_zh) )

    print('inserted: {} rows'.format(i))

    conn.commit()
    cur.close()
    conn.close()


if __name__ == "__main__":
    # parse_all()

    # load_file()

    tryMysql()

    print('done')
